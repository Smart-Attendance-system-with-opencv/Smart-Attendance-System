from tkinter import *
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from pytesseract import *
from openpyxl import *
import cv2
import re
# function for data extraction from image to list
def data_get_and_add():
   if((len(v.get())==0) or (len(w.get())==0) or (len(q.get())==0)):
          messagebox.showerror("Files Missing","Choose Files or Enter Date!!!")
   else:
        r = "016"
        img = cv2.imread(v.get())
        d = pytesseract.image_to_data(img, output_type=Output.DICT)
        l = list(d["text"])
        k = []
        nk=[]
        for i in l:
            if (re.match(r, i)):
                k.append(i)
        for i in k:
              nj=i.replace(i[0],"O",1)
              nk.append(nj)
        #excel
        wb = load_workbook(w.get())
        sheet = wb.active
        # add date and students attendance list to file
        dt = q.get()
        if(sheet.cell(row = 2 , column =1).value == None):
        	        messagebox.showerror("Missing","No Student List")
        else:	        
               
               x = 1;
               while(x<32):
               	    if(sheet.cell(row = 1, column = x). value == None):
               	    	  zx=x;
               	    	  break;
               	    x=x+1
               sheet.cell(row = 1 , column =zx).value = dt
               for i in nk:
               	    for o in range(2,92):
               	    	if(sheet.cell(row=o,column=1).value == i):
                             sheet.cell(row = o,column = zx).value = "Present"

               wb.save(w.get())
               messagebox.showinfo("Successfull","Check Your File for Attendance details...")
#add students id to file
def add_studentslist():
    if((len(v.get())==0) or (len(w.get())==0)):
        messagebox.showerror("Files Missing","Choose Files!!!")
    else:
        r = "016"
        img = cv2.imread(v.get())
        d = pytesseract.image_to_data(img, output_type=Output.DICT)
        l = list(d["text"])
        k = []
        nlk=[]
        for i in l:
            if (re.match(r, i)):
                k.append(i)
        for i in k:
            if (i != '0161207'):
                k.append('0161207')
                break;
        for i in k:
            if (i != '0161828'):
                k.append('0161828')
                break;
        for i in k:
               ns=i.replace(i[0],"O",1)
               nlk.append(ns)
        wb = load_workbook(w.get())
        sheet = wb.active
        h=2
        l=1
        sheet.cell(row=1, column=1).value = "No\Date"
        if(sheet.cell(row=2,column=1).value == None):
                for i in nlk:
                    sheet.cell(row=h,column=l).value = i
                    h=h+1
                    wb.save(w.get())
                messagebox.showinfo("Successfull", "Student Details filled Succesfully...")
        else:
             messagebox.showerror("File","Data filled already")
#screenshot
def browse_image():
     fn = filedialog.askopenfilename(initialdir="/home",filetypes=(("Image files","*.img"),("Image files","*.jpg"),("Image files","*.jpeg"),("All files","*.*")))
     e2.configure(text=fn)
     v.set(fn)
#file
def browse_file():
     f = filedialog.askopenfilename(initialdir="/home",filetypes=(("Excel files","*.xlsx"),("All files","*.*")))
     e3.configure(text = f)
     w.set(f)
#tkinter
root =tk.Tk()
root.title("Smart Attendance System")
root.geometry("1000x1000")

window = Canvas(root,width= 700,height = 700)
window.create_rectangle(10,130,660,480,outline = "black",fill = "#92A5FF")

label1 = Label(root,text = "Welcome to Our Application",font = ('Arial',18,'bold'))
label1.config(anchor = 'center')
label1.pack()
label2 = Label(root,text = "Working:",font=('Arial',15,'bold')).place( x = 100 , y =70)
label3 = Label(root,text = "1.Add Students List to Excel file").place( x = 130, y = 100)
label4 = Label(root,text = "2.Add Attendance Of Students to Excel File").place( x =130 , y = 120)

l1=Label(root, text ="Date",relief = "solid").place(x=200,y=200)
l2=Label(root, text ="Screenshot",relief = "solid").place(x=200,y=250)
l3=Label(root,text ="Choose File",relief = "solid").place(x=200 ,y=300)

q=tk.StringVar()
v=tk.StringVar()
w=tk.StringVar()

e1=Entry(root,width=45,textvariable =q).place(x= 300,y= 200)
e2=Label(root,width=45,height=1,textvariable=v,bg="white")
e2.place(x=300,y=250)
e3=Label(root,width=45,height=1,bg="white")
e3.place(x=300,y=300)

b1= Button(root,text = "Attendance"  ,    command=data_get_and_add, bd = 2,activebackground = "#62B1FF", bg = "#51DF6D").place(x=550,y=350)
b2= Button(root,text = "Choose Image",    command=browse_image    , bd = 2,activebackground = "#34F6A8", bg = "#EEE7B1").place(x=680,y=250)
b3= Button(root,text = "Select File" ,    command=browse_file     , bd = 2,activebackground = "#34F6A8", bg = "#EEE7B1").place(x=680,y=300)
b4= Button(root,text = "Student-List",    command=add_studentslist, bd = 2,activebackground = "#62B1FF", bg = "#51DF6D").place(x=300,y=350)
b5= Button(root,text = "Exit"        ,    command=exit            , bd = 2,activebackground = "red"    , bg = "#62B1FF").place(x=200,y=400)
window.pack()
root.resizable(0,0)
root.mainloop()